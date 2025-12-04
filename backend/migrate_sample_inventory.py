"""
Migration script to import data from sample_inventory.xlsx
This replaces the old migration with the new organized inventory format
"""
import openpyxl
from sqlalchemy.orm import Session
from app.core.database import SessionLocal, engine
from app.models.base import Base
from app.models.item import Item, Category
from app.models.location import Location, LocationType
from app.models.par_level import ParLevel
from app.models.user import User
from app.core.security import get_password_hash


def clear_database(db: Session):
    """Clear all existing data"""
    print("Clearing existing data...")
    db.query(ParLevel).delete()
    db.query(Item).delete()
    db.query(Category).delete()
    db.query(Location).delete()
    db.query(User).delete()
    db.commit()
    print("✓ Database cleared")


def create_locations(db: Session):
    """Create location hierarchy"""
    print("\nCreating locations...")
    
    # Main supply station
    main_station = Location(
        name="Main Supply Station",
        type=LocationType.SUPPLY_STATION,
        parent_location_id=None
    )
    db.add(main_station)
    db.flush()
    
    # Station cabinet
    station_cabinet = Location(
        name="Station Cabinet",
        type=LocationType.STATION_CABINET,
        parent_location_id=main_station.id
    )
    db.add(station_cabinet)
    db.flush()
    
    # Vehicles
    vehicles = []
    for i in range(1, 6):  # Create Medic 1-5
        vehicle = Location(
            name=f"Medic {i}",
            type=LocationType.VEHICLE,
            parent_location_id=station_cabinet.id
        )
        db.add(vehicle)
        vehicles.append(vehicle)
    
    db.flush()
    print(f"✓ Created {2 + len(vehicles)} locations")
    
    return {
        "main_station": main_station,
        "station_cabinet": station_cabinet,
        "vehicles": vehicles
    }


def create_categories(db: Session):
    """Create categories based on sheets"""
    print("\nCreating categories...")
    
    medical = Category(name="Medical Supplies")
    db.add(medical)
    
    cleaning = Category(name="Cleaning Supplies")
    db.add(cleaning)
    
    db.flush()
    print("✓ Created 2 categories")
    
    return {
        "medical": medical,
        "cleaning": cleaning
    }


def import_items_from_sheet(wb, sheet_name, category, db: Session, locations):
    """Import items from a specific sheet"""
    ws = wb[sheet_name]
    items_created = 0
    par_levels_created = 0
    seen_item_codes = set()
    
    # Skip header row, start from row 2
    for row_num in range(2, ws.max_row + 1):
        row = ws[row_num]
        
        # Get values
        item_id = row[0].value
        item_desc = row[1].value
        alt_location = row[2].value if len(row) > 2 else None
        par = row[3].value if len(row) > 3 else None
        reorder_at = row[4].value if len(row) > 4 else None
        uom = row[5].value if len(row) > 5 else "EA"
        
        # Skip if no item_id or description
        if not item_id or not item_desc:
            continue
        
        # Handle duplicate item codes by adding a suffix
        item_code = str(item_id).strip()
        original_code = item_code
        suffix = 1
        while item_code in seen_item_codes:
            item_code = f"{original_code}_{suffix}"
            suffix += 1
        seen_item_codes.add(item_code)
        
        # Create item
        item = Item(
            item_code=item_code,
            name=str(item_desc).strip(),
            description=f"Alternate Location: {alt_location}" if alt_location else None,
            category_id=category.id,
            unit_of_measure=str(uom).strip() if uom else "EA",
            is_active=True
        )
        db.add(item)
        db.flush()
        items_created += 1
        
        # Create par level for station cabinet if par value exists
        if par:
            try:
                par_int = int(float(par))  # Handle strings and decimals
                if par_int > 0:
                    # Get reorder quantity
                    reorder_int = int(par_int * 0.3)  # Default to 30% of par
                    if reorder_at:
                        try:
                            reorder_int = int(float(reorder_at))
                        except (ValueError, TypeError):
                            pass
                    
                    par_level = ParLevel(
                        item_id=item.id,
                        location_id=locations["station_cabinet"].id,
                        par_quantity=par_int,
                        reorder_quantity=reorder_int
                    )
                    db.add(par_level)
                    par_levels_created += 1
            except (ValueError, TypeError):
                pass  # Skip invalid par values
    
    return items_created, par_levels_created


def create_admin_user(db: Session):
    """Create admin user"""
    print("\nCreating admin user...")
    
    from app.models.user import UserRole
    
    admin = User(
        username="admin",
        email="admin@emstrack.local",
        password_hash=get_password_hash("ChangeMe123!"),
        first_name="System",
        last_name="Administrator",
        role=UserRole.ADMIN,
        is_active=True
    )
    db.add(admin)
    db.commit()
    print("✓ Admin user created (username: admin, password: ChangeMe123!)")


def main():
    """Main migration function"""
    print("=" * 60)
    print("EMS Supply Tracking - Sample Inventory Migration")
    print("=" * 60)
    
    # Create tables
    print("\nCreating database tables...")
    Base.metadata.create_all(bind=engine)
    print("✓ Tables created")
    
    # Load Excel file
    print("\nLoading sample_inventory.xlsx...")
    wb = openpyxl.load_workbook('../sample_inventory.xlsx')
    print(f"✓ Excel file loaded with sheets: {wb.sheetnames}")
    
    # Create database session
    db = SessionLocal()
    
    try:
        # Clear existing data
        clear_database(db)
        
        # Create locations
        locations = create_locations(db)
        
        # Create categories
        categories = create_categories(db)
        
        # Import medical supplies
        print("\nImporting medical supplies...")
        med_items, med_pars = import_items_from_sheet(
            wb, 
            "station_cabinet_medical", 
            categories["medical"], 
            db, 
            locations
        )
        print(f"✓ Created {med_items} medical items with {med_pars} par levels")
        
        # Import cleaning supplies
        print("\nImporting cleaning supplies...")
        clean_items, clean_pars = import_items_from_sheet(
            wb, 
            "station_cleaning_supplies", 
            categories["cleaning"], 
            db, 
            locations
        )
        print(f"✓ Created {clean_items} cleaning items with {clean_pars} par levels")
        
        # Create admin user
        create_admin_user(db)
        
        db.commit()
        
        print("\n" + "=" * 60)
        print("MIGRATION COMPLETED SUCCESSFULLY!")
        print("=" * 60)
        print(f"Total Items: {med_items + clean_items}")
        print(f"Total Par Levels: {med_pars + clean_pars}")
        print(f"Categories: {len(categories)}")
        print(f"Locations: {len(locations) + len(locations['vehicles'])}")
        print("\nLogin credentials:")
        print("  Username: admin")
        print("  Password: ChangeMe123!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n❌ Error during migration: {e}")
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
