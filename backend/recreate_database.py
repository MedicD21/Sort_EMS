"""
Recreate database with current schema and re-import sample data
"""
import sys
import os
from pathlib import Path

# Add backend directory to path
backend_dir = Path(__file__).parent
sys.path.insert(0, str(backend_dir))

from app.core.database import SessionLocal, engine, Base
from app.models.user import User
from app.models.location import Location, LocationType
from app.models.item import Category, Item
from app.models.par_level import ParLevel
from app.core.security import get_password_hash
import openpyxl
from sqlalchemy.orm import Session
from datetime import datetime


def recreate_database():
    """Drop all tables and recreate with current schema"""
    print("=" * 70)
    print("RECREATING DATABASE WITH CURRENT SCHEMA")
    print("=" * 70)
    
    # Drop all tables
    print("\n1. Dropping all existing tables...")
    Base.metadata.drop_all(bind=engine)
    print("   ✓ All tables dropped")
    
    # Create all tables with current schema
    print("\n2. Creating tables with current schema...")
    Base.metadata.create_all(bind=engine)
    print("   ✓ All tables created")


def create_admin_user(db: Session):
    """Create admin user"""
    print("\n3. Creating admin user...")
    admin = User(
        username="admin",
        email="admin@example.com",
        first_name="System",
        last_name="Administrator",
        password_hash=get_password_hash("Admin123!"),
        role="admin",
        is_active=True
    )
    db.add(admin)
    db.commit()
    print("   ✓ Admin user created (username: admin, password: Admin123!)")


def create_categories(db: Session):
    """Create standard EMS categories"""
    print("\n4. Creating categories...")
    
    categories_data = [
        ("AIRWAY", "Airway Management", "Airway devices, intubation equipment, suction", "#2196F3", 1),
        ("BREATHING", "Breathing & Oxygen", "Oxygen equipment, ventilation devices, nebulizers", "#03A9F4", 2),
        ("CARDIAC", "Cardiac Care", "Defibrillators, ECG equipment, cardiac medications", "#F44336", 3),
        ("TRAUMA", "Trauma Supplies", "Bandages, splints, bleeding control, c-collars", "#FF5722", 4),
        ("IV_FLUIDS", "IV & Fluids", "IV catheters, fluids, tubings, pumps", "#9C27B0", 5),
        ("MEDICATIONS", "Medications", "Emergency drugs, narcotics, controlled substances", "#4CAF50", 6),
        ("DIAGNOSTIC", "Diagnostic Equipment", "Blood pressure, pulse oximeter, thermometers, glucometers", "#FF9800", 7),
        ("INFECTION_CONTROL", "Infection Control", "PPE, gloves, masks, hand sanitizer, disinfectants", "#00BCD4", 8),
        ("OB_GYN", "OB/GYN", "Obstetric and gynecological supplies", "#E91E63", 9),
        ("PEDIATRIC", "Pediatric Equipment", "Pediatric-specific supplies and equipment", "#FFC107", 10),
        ("BURNS", "Burn Care", "Burn dressings, cooling gels, specialized burn supplies", "#FF6F00", 11),
        ("EXTRACTION", "Extraction & Rescue", "Extrication equipment, rescue tools, backboards", "#795548", 12),
        ("COMMUNICATION", "Communication", "Radios, phones, alert devices", "#607D8B", 13),
        ("DOCUMENTATION", "Documentation", "Forms, tablets, printers, pens", "#9E9E9E", 14),
        ("GENERAL", "General Supplies", "Miscellaneous supplies and equipment", "#757575", 99),
    ]
    
    for cat_id, name, desc, color, sort_order in categories_data:
        category = Category(
            id=cat_id,
            name=name,
            description=desc,
            color=color,
            sort_order=sort_order,
            is_active=True
        )
        db.add(category)
    
    db.commit()
    print(f"   ✓ Created {len(categories_data)} categories")
    return {cat[0]: cat[0] for cat in categories_data}


def create_locations(db: Session):
    """Create location hierarchy"""
    print("\n5. Creating locations...")
    
    # Main supply station
    main_station = Location(
        name="Main Supply Station",
        type=LocationType.SUPPLY_STATION,
        parent_location_id=None
    )
    db.add(main_station)
    db.flush()
    
    # Station 1 cabinet
    station_cabinet = Location(
        name="Station 1 Cabinet",
        type=LocationType.STATION_CABINET,
        parent_location_id=main_station.id
    )
    db.add(station_cabinet)
    db.flush()
    
    # Medic 4 vehicle
    medic4 = Location(
        name="Medic 4",
        type=LocationType.VEHICLE,
        parent_location_id=station_cabinet.id
    )
    db.add(medic4)
    db.flush()
    
    db.commit()
    print(f"   ✓ Created 3 locations")
    
    return {
        "main_station": main_station,
        "station_cabinet": station_cabinet,
        "medic4": medic4
    }


def import_items_from_excel(db: Session, categories, locations):
    """Import items from sample_inventory.xlsx"""
    print("\n6. Importing items from sample_inventory.xlsx...")
    
    excel_file = Path(__file__).parent.parent / "sample_inventory.xlsx"
    if not excel_file.exists():
        print(f"   ⚠ Excel file not found at {excel_file}")
        print("   Skipping item import")
        return
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active  # Use first sheet
    
    items_created = 0
    par_levels_created = 0
    
    # Mapping of keywords to categories
    category_keywords = {
        "TRAUMA": ["bandage", "gauze", "tape", "kling", "wrap", "dressing", "tourniquet"],
        "AIRWAY": ["airway", "intubat", "tube", "oxygen", "o2", "mask"],
        "IV_FLUIDS": ["iv", "saline", "catheter", "needle", "fluid"],
        "CARDIAC": ["cardiac", "defib", "ecg", "aed"],
        "MEDICATIONS": ["medication", "drug", "narcotic"],
        "DIAGNOSTIC": ["glucose", "blood pressure", "thermometer", "stethoscope"],
        "INFECTION_CONTROL": ["glove", "mask", "sanitizer", "ppe", "disinfect"],
        "BURNS": ["burn"],
        "GENERAL": []  # Default fallback
    }
    
    def categorize_item(item_name):
        """Determine category based on item name"""
        item_lower = item_name.lower()
        for category_id, keywords in category_keywords.items():
            if category_id == "GENERAL":
                continue
            for keyword in keywords:
                if keyword in item_lower:
                    return category_id
        return "GENERAL"  # Default
    
    # Skip header row
    for row_num in range(2, ws.max_row + 1):
        row = ws[row_num]
        
        item_code = str(row[0].value or "").strip()
        item_name = str(row[1].value or "").strip()
        
        if not item_code or not item_name:
            continue
        
        # Determine category
        category_id = categorize_item(item_name)
        
        # Create item
        item = Item(
            item_code=item_code,
            name=item_name,
            description=item_name,
            category_id=category_id,
            unit_of_measure="EA",
            is_active=True
        )
        db.add(item)
        items_created += 1
        
        # Create par level for Medic 4 (default to 1)
        db.flush()  # Get item.id
        par_level = ParLevel(
            item_id=item.id,
            location_id=locations["medic4"].id,
            par_level=1,
            is_active=True
        )
        db.add(par_level)
        par_levels_created += 1
    
    db.commit()
    print(f"   ✓ Created {items_created} items")
    print(f"   ✓ Created {par_levels_created} par levels")


def main():
    """Main migration function"""
    try:
        # Recreate database
        recreate_database()
        
        # Seed data
        db = SessionLocal()
        try:
            create_admin_user(db)
            categories = create_categories(db)
            locations = create_locations(db)
            import_items_from_excel(db, categories, locations)
            
            print("\n" + "=" * 70)
            print("✅ DATABASE RECREATION COMPLETE!")
            print("=" * 70)
            print("\nYou can now:")
            print("  - Login with: admin / Admin123!")
            print("  - View items in inventory management")
            print("  - Create new items with categories")
            print("  - Manage categories")
            
        finally:
            db.close()
            
    except Exception as e:
        print(f"\n❌ Error: {e}")
        raise


if __name__ == "__main__":
    main()
